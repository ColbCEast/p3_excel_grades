# Merrick Morgan, Colby Eastmond, Spencer Jorgensen, Jake Lee, Talon Condie
# This is a program that formats data within an excel file and cleans it

# Import openpyxl, workbook and font
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

# Load in the workbook
wb = openpyxl.load_workbook("Poorly_Organized_Data_1.xlsx")

# Set current sheet
current_sheet = wb.active

# Find values for each header
header = [cell.value for cell in current_sheet[1]]

# Find the column index for the column "Class Name"
if "Class Name" in header:
    class_idx = header.index("Class Name")
else:
    raise ValueError("Column 'Class Name' not found in the header")

# Find the column index for the column "Student Info"
if "Student Info" in header:
    student_idx = header.index("Student Info")
else:
    raise ValueError("Column 'Student Info' not found in the header")

# Find the column index for the column "Grade"
if "Grade" in header:
    grade_idx = header.index("Grade")
else:
    raise ValueError("Column 'Student Info' not found in the header")

# Initialize dictionary for students and classes
classes_and_students = {}

# Loop through the data, create a key-value pair for the classes and student information for each student taking the class
for row in current_sheet.iter_rows(min_row=2, values_only=True):
    class_name = row[class_idx]
    student_info = row[student_idx].split(sep = "_")
    grade = row[grade_idx]

    student_info.append(grade)

    if class_name in classes_and_students:
        classes_and_students[class_name].append(student_info)
    else:
        classes_and_students[class_name] = [student_info]

# Create a new workbook to work with
clean_workbook = Workbook()

# Pull the data from the excel sheet, clean it, and then format it
for key, value in classes_and_students.items():
    if key in clean_workbook.sheetnames:
        continue
    else:
        clean_workbook.create_sheet(key)

        worksheet = clean_workbook[key]
        
        headers = ['Last Name', 'First Name', 'Student ID', 'Grade']

        worksheet.append(headers)

        # Apply bold font to headers
        for col_num, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=col_num).font = Font(bold=True) 

        for i in range(len(value)):
            worksheet.append(value[i])

        for col_num, header in enumerate(headers, start=1):
            col_letter = worksheet.cell(row=1, column=col_num).column_letter  # Get column letter
            worksheet.column_dimensions[col_letter].width = len(header) + 5
        
        # Add a filter
        worksheet.auto_filter.ref = f"A1:D{worksheet.max_row}"

        # Get the Number of Students
        num_students = len(value)

        # information on the side with Summary grade info
        summary_info = ["Summary Statistics", "Highest Grade","Lowest Grade","Mean Grade","Median Grade","Number of Students"]
        summary_formulas = ["Value",
            f"=MAX(D2:D{worksheet.max_row})",
            f"=MIN(D2:D{worksheet.max_row})",
            f"=AVERAGE(D2:D{worksheet.max_row})",
            f"=MEDIAN(D2:D{worksheet.max_row})",
            f"=COUNTA(D2:D{worksheet.max_row})"]
        
        # Add the formatting to the names and the columns
        for i, (title, formula) in enumerate(zip(summary_info, summary_formulas), start=1):
            if i == 1:
                worksheet.cell(row=i, column=6, value=title).font = Font(bold=True)
                worksheet.cell(row=i, column=7, value=formula).font = Font(bold = True)
                worksheet.column_dimensions["F"].width = len(title) + 5
                worksheet.column_dimensions["G"].width = len(formula) + 5
            if i != 1:
                worksheet.cell(row=i, column=6, value=title)
                worksheet.cell(row=i, column=7, value=formula)


# Deleting the initial worksheet that is created
clean_workbook.remove(clean_workbook["Sheet"])

# Save the workbook
clean_workbook.save("formatted_grades.xlsx")

clean_workbook.close()





